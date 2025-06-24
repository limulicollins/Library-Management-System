from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QHBoxLayout, QPushButton, QComboBox,
    QLineEdit, QTableWidget, QTableWidgetItem, QDateEdit
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QIcon
from style import shared_stylesheet
from db_config import get_connection
from PyQt5.QtWidgets import QFrame
from PyQt5.QtGui import QPixmap
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

class ReportsPage(QWidget):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(shared_stylesheet)
        self.init_ui()

    def init_ui(self):

        main_layout = QVBoxLayout()
        title = QLabel("📄 Reports")
        title.setFont(QFont("Segoe UI", 20))
        title.setAlignment(Qt.AlignLeft)
        main_layout.addWidget(title)

        # --- Summary Cards Layout ---
        self.cards_layout = QHBoxLayout()
        self.cards_layout.setSpacing(20)

        # Initialize card widgets
        self.card1 = self.create_summary_card("Total", "0", "assets/icons/total.png")
        self.card2 = self.create_summary_card("Sub A", "0", "assets/icons/up.png")
        self.card3 = self.create_summary_card("Sub B", "0", "assets/icons/down.png")


        self.cards_layout.addWidget(self.card1)
        self.cards_layout.addWidget(self.card2)
        self.cards_layout.addWidget(self.card3)

        main_layout.addLayout(self.cards_layout)


        # Filter bar
        filter_layout = QHBoxLayout()
        self.report_type = QComboBox()
        self.report_type.addItems(["Books Report", "Members Report", "Transactions Report"])
        self.date_from = QDateEdit()
        self.date_to = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to.setDate(QDate.currentDate())

        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText("Search by keyword...")

        apply_btn = QPushButton("Apply")
        reset_btn = QPushButton("Reset")

        apply_btn.clicked.connect(self.load_report_data)
        self.report_type.currentTextChanged.connect(self.load_report_data)

        filter_layout.addWidget(QLabel("Report Type:"))
        filter_layout.addWidget(self.report_type)
        filter_layout.addWidget(QLabel("From:"))
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(QLabel("To:"))
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(self.search_field)
        filter_layout.addWidget(apply_btn)
        filter_layout.addWidget(reset_btn)

        main_layout.addLayout(filter_layout)

        # Table
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(5)
        self.report_table.setHorizontalHeaderLabels(["Column 1", "Column 2", "Column 3", "Column 4", "Column 5"])
        main_layout.addWidget(self.report_table)

        # Chart canvas
        self.chart_canvas = FigureCanvas(Figure(figsize=(5, 3)))
        main_layout.addWidget(self.chart_canvas)

        # Export buttons
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        export_pdf = QPushButton("Export PDF")
        export_excel = QPushButton("Export Excel")
        export_layout.addWidget(export_pdf)
        export_layout.addWidget(export_excel)

        export_pdf.clicked.connect(self.export_to_pdf)
        export_excel.clicked.connect(self.export_to_excel)

        main_layout.addLayout(export_layout)

        self.setLayout(main_layout)

    def reset_filters(self):
        self.search_field.clear()
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to.setDate(QDate.currentDate())
        self.load_report_data()


    def load_report_data(self):
        report_type = self.report_type.currentText()
        search_text = self.search_field.text().lower()
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")

        try:
            conn = get_connection()
            cursor = conn.cursor()

            if report_type == "Books Report":
                query = """
                    SELECT id, author, isbn, genre, copies 
                    FROM books
                    WHERE LOWER(author) LIKE %s OR LOWER(genre) LIKE %s OR LOWER(isbn) LIKE %s
                """
                cursor.execute(query, (f'%{search_text}%', f'%{search_text}%', f'%{search_text}%'))
                results = cursor.fetchall()
                headers = ["ID", "Author", "ISBN", "Genre", "Copies"]

            elif report_type == "Members Report":
                query = """
                    SELECT id, name, email, phone, address 
                    FROM members
                    WHERE LOWER(name) LIKE %s OR LOWER(email) LIKE %s
                """
                cursor.execute(query, (f'%{search_text}%', f'%{search_text}%'))
                results = cursor.fetchall()
                headers = ["ID", "Name", "Email", "Phone", "Address"]

            elif report_type == "Transactions Report":
                query = """
                    SELECT t.id, b.title, m.name, t.issue_date, t.return_date, t.returned
                    FROM transactions t
                    JOIN books b ON t.book_id = b.id
                    JOIN members m ON t.member_id = m.id
                    WHERE (t.issue_date BETWEEN %s AND %s OR t.return_date BETWEEN %s AND %s)
                    AND (LOWER(b.title) LIKE %s OR LOWER(m.name) LIKE %s)
                """
                cursor.execute(query, (
                    date_from, date_to, date_from, date_to,
                    f'%{search_text}%', f'%{search_text}%'
                ))
                results = cursor.fetchall()
                headers = ["ID", "Book", "Member", "Issue Date", "Return Date", "Returned"]

            self.populate_table(results, headers)
            conn.close()

            # Update summary cards
            if report_type == "Books Report":
                total = len(results)
                most = results[0][1] if results else "N/A"
                least = results[-1][1] if results else "N/A"
                self.update_cards("Total Books", total, "Top Author", most, "Bottom Author", least)

            elif report_type == "Members Report":
                total = len(results)
                active = sum(1 for row in results if "active" in str(row).lower())
                inactive = total - active
                self.update_cards("Total Members", total, "Active", active, "Inactive", inactive)

            elif report_type == "Transactions Report":
                total = len(results)
                returned = sum(1 for row in results if str(row[-1]).lower() == "yes")
                overdue = total - returned
                self.update_cards("Total Txns", total, "Returned", returned, "Overdue", overdue)

            self.update_chart(report_type, results)

        except Exception as e:
            print("Error loading report:", e)

        

    def populate_table(self, data, headers):
        self.report_table.clear()
        self.report_table.setRowCount(len(data))
        self.report_table.setColumnCount(len(headers))
        self.report_table.setHorizontalHeaderLabels(headers)

        for row_idx, row_data in enumerate(data):
            for col_idx, item in enumerate(row_data):
                self.report_table.setItem(row_idx, col_idx, QTableWidgetItem(str(item)))

    def export_to_excel(self):
        from openpyxl import Workbook
        from PyQt5.QtWidgets import QFileDialog, QMessageBox

        path, _ = QFileDialog.getSaveFileName(self, "Save Report", "", "Excel Files (*.xlsx)")
        if not path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"

        # Headers
        for col in range(self.report_table.columnCount()):
            sheet.cell(row=1, column=col+1).value = self.report_table.horizontalHeaderItem(col).text()

        # Data
        for row in range(self.report_table.rowCount()):
            for col in range(self.report_table.columnCount()):
                item = self.report_table.item(row, col)
                sheet.cell(row=row+2, column=col+1).value = item.text() if item else ""

        workbook.save(path)
        QMessageBox.information(self, "Exported", "Report successfully exported to Excel.")

    def export_to_pdf(self):
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from PyQt5.QtWidgets import QFileDialog, QMessageBox

        path, _ = QFileDialog.getSaveFileName(self, "Save Report", "", "PDF Files (*.pdf)")
        if not path:
            return

        c = canvas.Canvas(path, pagesize=letter)
        width, height = letter
        x_offset, y_offset = 50, height - 50
        row_height = 20

        # Headers
        for col in range(self.report_table.columnCount()):
            text = self.report_table.horizontalHeaderItem(col).text()
            c.drawString(x_offset + col*100, y_offset, text)

        # Data
        y_offset -= row_height
        for row in range(self.report_table.rowCount()):
            for col in range(self.report_table.columnCount()):
                item = self.report_table.item(row, col)
                c.drawString(x_offset + col*100, y_offset, item.text() if item else "")
            y_offset -= row_height
            if y_offset < 50:
                c.showPage()
                y_offset = height - 50

        c.save()
        QMessageBox.information(self, "Exported", "Report successfully exported to PDF.")

    def create_summary_card(self, title, value, icon_path=None):
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                border-radius: 10px;
                padding: 10px;
            }
            QLabel {
                color: white;
            }
        """)
        card_layout = QHBoxLayout()
    
        if icon_path:
            icon_label = QLabel()
            pixmap = QPixmap(icon_path).scaled(40, 40, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            icon_label.setPixmap(pixmap)
            card_layout.addWidget(icon_label)

        text_layout = QVBoxLayout()
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 14px; color: #bbb;")
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #4fd1c5;")
        text_layout.addWidget(title_label)
        text_layout.addWidget(value_label)

        card_layout.addLayout(text_layout)
        card.setLayout(card_layout)
        return card

    
    def update_cards(self, title1, val1, title2, val2, title3, val3):
        self.update_card(0, title1, val1)
        self.update_card(1, title2, val2)
        self.update_card(2, title3, val3)

    def update_card(self, index, title, value):
        card = self.cards_layout.itemAt(index).widget()
        text_layout = card.layout().itemAt(1)
        title_label = text_layout.layout().itemAt(0).widget()
        value_label = text_layout.layout().itemAt(1).widget()
        title_label.setText(title)
        value_label.setText(str(value))

    def update_chart(self, report_type, data):
        self.chart_canvas.figure.clear()
        ax = self.chart_canvas.figure.add_subplot(111)

        if report_type == "Books Report":
            if data:
                authors = [row[1] for row in data[:5]]
                copies = [int(row[4]) for row in data[:5]]
                ax.bar(authors, copies, color='#4fd1c5')
                ax.set_title("Top 5 Authors by Copies")

        elif report_type == "Members Report":
            total = len(data)
            active = sum(1 for row in data if "active" in str(row).lower())
            inactive = total - active
            ax.pie([active, inactive], labels=["Active", "Inactive"], autopct='%1.1f%%', colors=['#4fd1c5', '#fd6e6e'])
            ax.set_title("Member Status Distribution")

        elif report_type == "Transactions Report":
            # Count transactions per day
            from collections import Counter
            import datetime

            dates = [str(row[3]) for row in data if row[3]]  # issue_date
            counts = Counter(dates)
            sorted_dates = sorted(counts)
            values = [counts[d] for d in sorted_dates]

            ax.plot(sorted_dates, values, marker='o', color='#facc15')
            ax.set_title("Transactions Over Time")
            ax.tick_params(axis='x', rotation=45)

        self.chart_canvas.draw()






    
